#!/usr/bin/env python3
"""Safe maintenance helper for the llm-rl-visualized asset tree.

This helper mirrors the repo's source-script intent while adding dry-run-first
behavior, explicit apply flags, root selection, and collision checks.
"""

from __future__ import annotations

import argparse
import re
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

IMAGE_DIRS = {
    "chinese": "images_chinese",
    "english": "images_english",
}

WORKBOOKS = {
    "chinese": Path("src/conf/info-ch.xlsx"),
    "english": Path("src/conf/info-en.xlsx"),
}

SUPPORTED_IMAGE_SUFFIXES = {".png", ".svg"}
DEFAULT_PATTERN = r"^幻灯片(\d+)"


class CliError(RuntimeError):
    """Raised for user-facing CLI errors."""


@dataclass(frozen=True)
class WorkbookRow:
    row_index: int
    category: str
    title: str
    existing_name: str
    expected_name: str


def text(value) -> str:
    return "" if value is None else str(value)


def derive_name(category, title) -> str:
    category_text = text(category)
    title_text = text(title)
    if not category_text or not title_text:
        return ""
    return f"【{category_text}】{title_text}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Inspect, preview, and safely maintain the bilingual diagram asset tree."
    )
    parser.add_argument(
        "--root",
        "--repo",
        dest="root",
        type=Path,
        default=Path.cwd(),
        help="Repository root or copied checkout root. Defaults to the current directory.",
    )

    subparsers = parser.add_subparsers(dest="command", required=True)

    inventory = subparsers.add_parser(
        "inventory",
        help="Summarize asset counts, workbook rows, and tree drift.",
    )
    inventory.add_argument(
        "--language",
        choices=("all", "chinese", "english"),
        default="all",
        help="Limit the inventory to one language or inspect both trees.",
    )

    rename = subparsers.add_parser(
        "rename-plan",
        help="Preview or apply workbook-driven renames for slide exports.",
    )
    rename.add_argument(
        "--language",
        choices=("all", "chinese", "english"),
        default="all",
        help="Choose which language tree to scan.",
    )
    rename.add_argument(
        "--pattern",
        default=DEFAULT_PATTERN,
        help="Regex for slide-export names. The first capture group must be the slide number.",
    )
    rename.add_argument(
        "--apply",
        action="store_true",
        help="Apply the rename plan instead of previewing it.",
    )
    rename.add_argument(
        "--force",
        action="store_true",
        help="Allow the helper to replace an existing target path during apply.",
    )

    add_name = subparsers.add_parser(
        "add-name-column",
        help="Preview or refresh the workbook's generated name column.",
    )
    add_name.add_argument(
        "--language",
        choices=("all", "chinese", "english"),
        default="all",
        help="Choose which workbook(s) to update.",
    )
    add_name.add_argument(
        "--apply",
        action="store_true",
        help="Write the refreshed workbook(s) instead of previewing changes.",
    )
    add_name.add_argument(
        "--output-dir",
        type=Path,
        help="Optional output directory for copied workbook(s). If omitted, apply writes in place.",
    )
    add_name.add_argument(
        "--force",
        action="store_true",
        help="Allow existing output workbook(s) to be replaced.",
    )

    trim = subparsers.add_parser(
        "trim",
        help="Preview or apply whitespace trimming to PNG assets.",
    )
    trim.add_argument(
        "--language",
        choices=("all", "chinese", "english"),
        default="all",
        help="Choose which repo image tree to trim when --input-dir is not supplied.",
    )
    trim.add_argument(
        "--input-dir",
        type=Path,
        help="Optional custom input directory to trim instead of the repo defaults.",
    )
    trim.add_argument(
        "--output-dir",
        type=Path,
        help="Optional output directory for trimmed copies. Required with --apply.",
    )
    trim.add_argument(
        "--padding",
        type=int,
        default=4,
        help="Extra border to keep around cropped content. Defaults to the source-script padding.",
    )
    trim.add_argument(
        "--apply",
        action="store_true",
        help="Write trimmed PNGs to the output directory instead of previewing them.",
    )
    trim.add_argument(
        "--force",
        action="store_true",
        help="Allow existing output files to be replaced during apply.",
    )

    return parser.parse_args()


def resolve_root(root: Path) -> Path:
    resolved = root.expanduser().resolve()
    if not resolved.exists():
        raise CliError(f"root does not exist: {resolved}")
    return resolved


def selected_languages(choice: str) -> tuple[str, ...]:
    if choice == "all":
        return ("chinese", "english")
    return (choice,)


def require_openpyxl(command: str):
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - runtime guard
        raise CliError(
            f"{command} requires openpyxl to read the workbook files. Install openpyxl in the inspection environment and retry."
        ) from exc
    return openpyxl


def require_pillow(command: str):
    try:
        from PIL import Image, ImageChops
    except ImportError as exc:  # pragma: no cover - runtime guard
        raise CliError(
            f"{command} requires Pillow to trim PNG assets. Install Pillow in the inspection environment and retry."
        ) from exc
    return Image, ImageChops


def iter_files(root: Path, suffixes: Iterable[str]) -> list[Path]:
    suffix_set = {suffix.lower() for suffix in suffixes}
    files = [p for p in root.rglob("*") if p.is_file() and p.suffix.lower() in suffix_set]
    return sorted(files)


def read_workbook_rows(workbook_path: Path) -> list[WorkbookRow]:
    openpyxl = require_openpyxl("workbook commands")
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb.active
    rows: list[WorkbookRow] = []
    for row_index in range(1, ws.max_row + 1):
        category = text(ws.cell(row_index, 1).value)
        title = text(ws.cell(row_index, 2).value)
        existing_name = text(ws.cell(row_index, 3).value)
        expected_name = existing_name if existing_name else derive_name(category, title)
        if not existing_name and not category and not title:
            expected_name = ""
        rows.append(
            WorkbookRow(
                row_index=row_index,
                category=category,
                title=title,
                existing_name=existing_name,
                expected_name=expected_name,
            )
        )
    return rows


def load_workbook_and_sheet(workbook_path: Path):
    openpyxl = require_openpyxl("workbook commands")
    wb = openpyxl.load_workbook(workbook_path)
    return wb, wb.active


def rel_for_output(path: Path, base: Path) -> Path:
    try:
        return path.relative_to(base)
    except ValueError:
        return Path(path.name)


def within(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
        return True
    except ValueError:
        return False


def sample(items: Iterable[str], limit: int = 5) -> str:
    items = list(items)
    if not items:
        return "-"
    if len(items) <= limit:
        return ", ".join(items)
    return ", ".join(items[:limit]) + f" ... (+{len(items) - limit} more)"


def report_delta(label_left: str, left: set[str], label_right: str, right: set[str]) -> None:
    only_left = sorted(left - right)
    only_right = sorted(right - left)
    if not only_left and not only_right:
        print(f"    {label_left} vs {label_right}: aligned")
        return
    print(f"    {label_left} vs {label_right}:")
    if only_left:
        print(f"      only in {label_left} ({len(only_left)}): {sample(only_left)}")
    if only_right:
        print(f"      only in {label_right} ({len(only_right)}): {sample(only_right)}")


def collect_stems(root: Path, suffix: str) -> set[str]:
    if not root.exists():
        return set()
    return {p.stem for p in iter_files(root, (suffix,))}


def inventory_command(root: Path, language: str) -> int:
    print(f"root: {root}")
    langs = selected_languages(language)
    for lang in langs:
        image_root = root / IMAGE_DIRS[lang]
        workbook_path = root / WORKBOOKS[lang]
        print(f"\n[{lang}]")
        big = collect_stems(image_root / "png_big", ".png")
        small = collect_stems(image_root / "png_small", ".png")
        svg = collect_stems(image_root / "source_svg", ".svg")
        print(f"  png_big: {len(big)}")
        print(f"  png_small: {len(small)}")
        print(f"  source_svg: {len(svg)}")

        xlsx_dir = image_root / "source_xlsx"
        if xlsx_dir.exists():
            xlsx_files = sorted(p.name for p in xlsx_dir.glob("*.xlsx") if p.is_file())
            print(f"  source_xlsx: {len(xlsx_files)}")
            if xlsx_files:
                print(f"    files: {sample(xlsx_files, limit=10)}")

        if workbook_path.exists():
            try:
                rows = read_workbook_rows(workbook_path)
                workbook_stems = {row.expected_name for row in rows if row.expected_name}
                print(f"  workbook rows: {len(rows)}")
                report_delta("png_big", big, "png_small", small)
                report_delta("png_big", big, "source_svg", svg)
                report_delta("png_big", big, "workbook", workbook_stems)
            except CliError as exc:
                print(f"  workbook stats skipped: {exc}")
        else:
            print(f"  workbook missing: {workbook_path}")

    print("\n[assets]")
    asset_files = [
        root / "src/assets/Book_Cover.png",
        root / "src/assets/LLM-RL-Algorithms-en.png",
        root / "src/assets/WeChat.jpg",
        root / "src/assets/account.png",
        root / "src/assets/banner.pptx",
        root / "src/assets/book_url.png",
        root / "src/assets/images-template.pptx",
        root / "src/assets/大模型算法-内容架构-zh.png",
        root / "src/assets/banner/幻灯片1.SVG",
        root / "src/assets/banner/幻灯片2.SVG",
    ]
    present_assets = [p.relative_to(root).as_posix() for p in asset_files if p.exists()]
    print(f"  src/assets files: {len(present_assets)}")
    print(f"    files: {sample(present_assets, limit=10)}")

    pdf_files = [
        root / "强化学习算法图谱 (rl-algo-map).pdf",
        root / "策略梯度(Policy Gradient)-强化学习(PPO&GRPO等)之根基.pdf",
        root / "src/大模型与强化学习——算法架构、工程体系与前沿.pdf",
    ]
    present_pdfs = [p.relative_to(root).as_posix() for p in pdf_files if p.exists()]
    print(f"  pdf anchors: {len(present_pdfs)}")
    print(f"    files: {sample(present_pdfs, limit=10)}")
    return 0


def build_name_rows(workbook_path: Path) -> list[WorkbookRow]:
    return read_workbook_rows(workbook_path)


def rename_plan_for_language(root: Path, lang: str, pattern: re.Pattern[str], apply: bool, force: bool) -> bool:
    image_root = root / IMAGE_DIRS[lang]
    workbook_path = root / WORKBOOKS[lang]
    if not image_root.exists():
        print(f"[{lang}] image tree missing: {image_root}")
        return False
    if not workbook_path.exists():
        print(f"[{lang}] workbook missing: {workbook_path}")
        return False

    rows = build_name_rows(workbook_path)
    files = [p for p in iter_files(image_root, SUPPORTED_IMAGE_SUFFIXES) if pattern.match(p.name)]

    changed = False
    collisions = 0
    out_of_range = 0
    no_map = 0
    unchanged = 0

    for src in files:
        match = pattern.match(src.name)
        if not match:
            continue
        slide_num = int(match.group(1))
        if slide_num < 1 or slide_num > len(rows):
            out_of_range += 1
            print(
                f"[{lang}] {src.relative_to(root).as_posix()} -> no workbook row {slide_num} (workbook rows: {len(rows)})"
            )
            continue
        target_stem = rows[slide_num - 1].expected_name
        if not target_stem:
            no_map += 1
            print(f"[{lang}] {src.relative_to(root).as_posix()} -> missing target stem in workbook row {slide_num}")
            continue
        if any(sep in target_stem for sep in ("/", "\\")):
            raise CliError(f"workbook row {slide_num} produces an invalid file stem: {target_stem!r}")
        target = src.with_name(target_stem + src.suffix)
        rel_src = src.relative_to(root).as_posix()
        rel_target = target.relative_to(root).as_posix() if target.is_relative_to(root) else str(target)
        if target == src:
            unchanged += 1
            print(f"[{lang}] {rel_src} -> already named")
            continue
        if target.exists() and target != src and not force:
            collisions += 1
            print(f"[{lang}] {rel_src} -> collision at {rel_target} (use --force only after review)")
            continue
        changed = True
        if apply:
            target.parent.mkdir(parents=True, exist_ok=True)
            if force and target.exists() and target != src:
                if target.is_dir():
                    raise CliError(f"cannot replace directory target: {rel_target}")
                target.unlink()
            src.rename(target)
            print(f"[{lang}] renamed: {rel_src} -> {rel_target}")
        else:
            print(f"[{lang}] plan: {rel_src} -> {rel_target}")

    if not files:
        print(f"[{lang}] no matching slide-prefixed image files found under {image_root.relative_to(root).as_posix()}")

    if apply:
        if collisions or out_of_range or no_map:
            print(
                f"[{lang}] apply finished with warnings: collisions={collisions}, out_of_range={out_of_range}, missing_names={no_map}, unchanged={unchanged}"
            )
        else:
            print(f"[{lang}] apply finished cleanly: renamed={int(changed)}, unchanged={unchanged}")
    return changed or bool(collisions or out_of_range or no_map)


def rename_plan_command(root: Path, language: str, pattern_text: str, apply: bool, force: bool) -> int:
    pattern = re.compile(pattern_text)
    for lang in selected_languages(language):
        rename_plan_for_language(root, lang, pattern, apply, force)
    return 0


def refresh_name_column_for_language(
    root: Path,
    lang: str,
    apply: bool,
    output_dir: Path | None,
    force: bool,
) -> bool:
    workbook_path = root / WORKBOOKS[lang]
    if not workbook_path.exists():
        print(f"[{lang}] workbook missing: {workbook_path}")
        return False

    wb, ws = load_workbook_and_sheet(workbook_path)
    changed_rows: list[str] = []
    for row_index in range(1, ws.max_row + 1):
        category = text(ws.cell(row_index, 1).value)
        title = text(ws.cell(row_index, 2).value)
        existing = text(ws.cell(row_index, 3).value)
        derived = derive_name(category, title) if (category or title) else existing
        if derived != existing:
            changed_rows.append(f"row {row_index}: {existing!r} -> {derived!r}")
        if apply:
            ws.cell(row_index, 3).value = derived

    rel_workbook = workbook_path.relative_to(root).as_posix() if workbook_path.is_relative_to(root) else workbook_path.name
    if not apply:
        print(f"[{lang}] {rel_workbook}: {len(changed_rows)} row(s) would change")
        for line in changed_rows[:25]:
            print(f"  {line}")
        if len(changed_rows) > 25:
            print(f"  ... (+{len(changed_rows) - 25} more)")
        return bool(changed_rows)

    if output_dir is None:
        dest = workbook_path
    else:
        dest = output_dir / rel_workbook
        if dest.exists() and not force:
            print(f"[{lang}] output exists, skipping without --force: {dest.relative_to(output_dir).as_posix() if dest.is_relative_to(output_dir) else dest}")
            return False
        dest.parent.mkdir(parents=True, exist_ok=True)

    wb.save(dest)
    print(f"[{lang}] wrote workbook: {dest.relative_to(root).as_posix() if dest.is_relative_to(root) else dest}")
    if changed_rows:
        print(f"[{lang}] refreshed {len(changed_rows)} row(s)")
    else:
        print(f"[{lang}] no workbook row changed")
    return bool(changed_rows)


def add_name_column_command(root: Path, language: str, apply: bool, output_dir: Path | None, force: bool) -> int:
    for lang in selected_languages(language):
        refresh_name_column_for_language(root, lang, apply, output_dir, force)
    return 0


def flatten_on_white(Image, im):
    """Flatten an image against white before trimming.

    The source trim script assumes a white background; this helper preserves
    that behavior while staying robust for RGB, RGBA, and palette images.
    """
    if im.mode == "RGB":
        return im.copy()
    rgba = im.convert("RGBA")
    white = Image.new("RGBA", rgba.size, (255, 255, 255, 255))
    return Image.alpha_composite(white, rgba).convert("RGB")


def trim_single_image(
    Image,
    ImageChops,
    src: Path,
    dst: Path,
    padding: int,
) -> tuple[str, tuple[int, int, int, int] | None, tuple[int, int]]:
    with Image.open(src) as im:
        flat = flatten_on_white(Image, im)
        width, height = flat.size
        bg = Image.new("RGB", flat.size, (255, 255, 255))
        diff = ImageChops.difference(flat, bg)
        bbox = diff.getbbox()
        if bbox:
            left = max(bbox[0] - padding, 0)
            upper = max(bbox[1] - padding, 0)
            right = min(bbox[2] + padding, width)
            lower = min(bbox[3] + padding // 2, height)
            cropped = flat.crop((left, upper, right, lower))
            return "cropped", (left, upper, right, lower), cropped.size
        return "unchanged", None, flat.size


def trim_tree(
    Image,
    ImageChops,
    source_root: Path,
    base_root: Path,
    output_root: Path | None,
    padding: int,
    apply: bool,
    force: bool,
) -> bool:
    pngs = [p for p in iter_files(source_root, (".png",))]
    if not pngs:
        print(f"[trim] no PNG files found under {source_root.relative_to(base_root).as_posix() if source_root.is_relative_to(base_root) else source_root}")
        return False

    changed = False
    for src in pngs:
        rel = src.relative_to(base_root)
        dst = output_root / rel if output_root is not None else None
        status, bbox, size = trim_single_image(Image, ImageChops, src, dst if dst is not None else src, padding)
        rel_src = rel.as_posix()
        if status == "unchanged":
            if apply and dst is not None:
                if dst.exists() and not force:
                    print(f"[trim] skip existing output: {dst.relative_to(output_root).as_posix()}")
                    continue
                dst.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(src, dst)
            print(f"[trim] {rel_src} -> unchanged")
            continue

        assert bbox is not None
        changed = True
        if apply:
            if dst is None:
                raise CliError("trim --apply requires --output-dir")
            if dst.exists() and not force:
                print(f"[trim] skip existing output: {dst.relative_to(output_root).as_posix()}")
                continue
            dst.parent.mkdir(parents=True, exist_ok=True)
            with Image.open(src) as im:
                flat = flatten_on_white(Image, im)
                crop_left, crop_upper, crop_right, crop_lower = bbox
                flat.crop((crop_left, crop_upper, crop_right, crop_lower)).save(dst)
            print(f"[trim] wrote {rel_src} -> {dst.relative_to(output_root).as_posix()}")
        else:
            if dst is not None:
                rel_dst = dst.relative_to(output_root).as_posix() if output_root is not None else str(dst)
                print(f"[trim] plan: {rel_src} -> {rel_dst} bbox={bbox} size={size}")
            else:
                print(f"[trim] plan: {rel_src} bbox={bbox} size={size}")
    return changed


def trim_command(
    root: Path,
    language: str,
    input_dir: Path | None,
    output_dir: Path | None,
    padding: int,
    apply: bool,
    force: bool,
) -> int:
    Image, ImageChops = require_pillow("trim")

    if input_dir is not None:
        source_root = input_dir.expanduser().resolve()
        if not source_root.exists():
            raise CliError(f"input directory does not exist: {source_root}")
        sources = [(source_root, source_root)]
    else:
        sources = []
        for lang in selected_languages(language):
            source_root = root / IMAGE_DIRS[lang]
            if source_root.exists():
                sources.append((source_root, root))
            else:
                print(f"[trim] missing source tree: {source_root}")

    if apply and output_dir is None:
        raise CliError("trim --apply requires --output-dir so originals are never overwritten")

    output_root = output_dir.expanduser().resolve() if output_dir is not None else None
    if output_root is not None:
        for source_root, base_root in sources:
            if output_root == source_root or within(output_root, source_root):
                raise CliError("output-dir must be outside the input tree")

    if not sources:
        print("[trim] no input tree selected")
        return 0

    for source_root, base_root in sources:
        trim_tree(Image, ImageChops, source_root, base_root, output_root, padding, apply, force)
    return 0


def main() -> int:
    try:
        args = parse_args()
        root = resolve_root(args.root)
        if args.command == "inventory":
            return inventory_command(root, args.language)
        if args.command == "rename-plan":
            return rename_plan_command(root, args.language, args.pattern, args.apply, args.force)
        if args.command == "add-name-column":
            return add_name_column_command(root, args.language, args.apply, args.output_dir, args.force)
        if args.command == "trim":
            return trim_command(root, args.language, args.input_dir, args.output_dir, args.padding, args.apply, args.force)
        raise CliError(f"unknown command: {args.command}")
    except CliError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
