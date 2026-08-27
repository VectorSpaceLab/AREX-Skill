#!/usr/bin/env python3
"""Inspect Skywork evaluation output files without running benchmarks."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - optional dependency path
    load_workbook = None


COMMON_FIELDS = [
    "id",
    "question",
    "answer",
    "prediction",
    "response",
    "score",
    "hit",
    "error",
]


def _inspect_jsonl(path: Path) -> Dict[str, Any]:
    total = 0
    valid = 0
    invalid = 0
    keys = Counter()
    response_count = 0
    error_count = 0
    for line_num, line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        if not line.strip():
            continue
        total += 1
        try:
            record = json.loads(line)
        except json.JSONDecodeError:
            invalid += 1
            continue
        if not isinstance(record, dict):
            invalid += 1
            continue
        valid += 1
        keys.update(record.keys())
        if record.get("response") or record.get("prediction"):
            response_count += 1
        response_block = record.get("response")
        if isinstance(response_block, dict) and response_block.get("error"):
            error_count += 1
        elif record.get("error"):
            error_count += 1
    return {
        "type": "jsonl",
        "total_lines": total,
        "valid_records": valid,
        "invalid_records": invalid,
        "top_keys": keys.most_common(20),
        "records_with_response_or_prediction": response_count,
        "records_with_error": error_count,
    }


def _inspect_json(path: Path) -> Dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        values = list(data.values())
        dict_values = [v for v in values if isinstance(v, dict)]
        key_counter = Counter()
        for value in dict_values:
            key_counter.update(value.keys())
        return {
            "type": "json-object",
            "top_level_keys": len(data),
            "dict_value_records": len(dict_values),
            "top_record_keys": key_counter.most_common(20),
            "common_fields_present": [field for field in COMMON_FIELDS if any(field in r for r in dict_values)],
        }
    if isinstance(data, list):
        dict_records = [r for r in data if isinstance(r, dict)]
        key_counter = Counter()
        for record in dict_records:
            key_counter.update(record.keys())
        return {
            "type": "json-list",
            "records": len(data),
            "dict_records": len(dict_records),
            "top_record_keys": key_counter.most_common(20),
            "common_fields_present": [field for field in COMMON_FIELDS if any(field in r for r in dict_records)],
        }
    return {"type": "json-other", "python_type": type(data).__name__}


def _inspect_xlsx(path: Path) -> Dict[str, Any]:
    if load_workbook is None:
        return {
            "type": "xlsx",
            "warning": "openpyxl is not installed; only file existence was checked",
        }
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, ())
        header_text = ["" if h is None else str(h) for h in headers]
        sheets.append(
            {
                "name": sheet_name,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "headers": header_text[:30],
                "common_fields_present": [field for field in COMMON_FIELDS if field in header_text],
            }
        )
    return {"type": "xlsx", "sheets": sheets}


def inspect_path(path: str) -> Dict[str, Any]:
    source = Path(path)
    result: Dict[str, Any] = {
        "path": path,
        "exists": source.exists(),
    }
    if not source.exists():
        result["status"] = "missing"
        return result
    result["size_bytes"] = source.stat().st_size
    suffix = source.suffix.lower()
    try:
        if suffix == ".jsonl":
            result.update(_inspect_jsonl(source))
        elif suffix == ".json":
            result.update(_inspect_json(source))
        elif suffix in {".xlsx", ".xlsm"}:
            result.update(_inspect_xlsx(source))
        else:
            result.update({"type": suffix.lstrip(".") or "unknown"})
        result["status"] = "ok"
    except Exception as exc:
        result["status"] = "error"
        result["error"] = str(exc)
    return result


def _print_human(results: Iterable[Dict[str, Any]]) -> None:
    for result in results:
        print(f"Path: {result['path']}")
        print(f"  Exists: {result['exists']}")
        print(f"  Status: {result.get('status')}")
        if result.get("exists"):
            print(f"  Type: {result.get('type')}")
            print(f"  Size: {result.get('size_bytes')} bytes")
        if result.get("error"):
            print(f"  Error: {result['error']}")
        if result.get("top_keys"):
            print(f"  Top keys: {result['top_keys']}")
        if result.get("top_record_keys"):
            print(f"  Top record keys: {result['top_record_keys']}")
        if result.get("sheets"):
            print(f"  Sheets: {[sheet['name'] for sheet in result['sheets']]}")
        print()


def main() -> int:
    parser = argparse.ArgumentParser(description="Inspect Skywork evaluation output files without running benchmarks.")
    parser.add_argument("--input", nargs="+", required=True, help="One or more JSON, JSONL, or XLSX output files.")
    parser.add_argument("--json", action="store_true", help="Print JSON summary.")
    args = parser.parse_args()

    results = [inspect_path(path) for path in args.input]
    if args.json:
        print(json.dumps({"results": results}, ensure_ascii=False, indent=2))
    else:
        _print_human(results)
    return 0 if all(result.get("status") == "ok" for result in results) else 1


if __name__ == "__main__":
    raise SystemExit(main())
