#!/usr/bin/env python3
"""Score Skywork evaluation records with the last boxed answer.

The helper accepts JSON, JSONL, or XLSX input and writes annotated records plus
an accuracy summary without calling any remote judge.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - optional dependency path
    load_workbook = None

BOXED_RE = re.compile(r"\\boxed\{(.*?)\}", re.DOTALL)
TAG_RE = re.compile(r"<answer>(.*?)</answer>", re.DOTALL)


def _normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\$|\\[a-zA-Z]+|\s+", "", text)
    text = text.strip()
    if text.startswith("{") and text.endswith("}"):
        text = text[1:-1]
    return text.lower()


def _extract_prediction(record: Dict[str, Any]) -> str:
    for key in ("prediction", "response", "model_response", "output"):
        value = record.get(key)
        if isinstance(value, str) and value.strip():
            return value
        if isinstance(value, dict):
            for nested_key in ("content", "full_response", "answer", "text"):
                nested = value.get(nested_key)
                if isinstance(nested, str) and nested.strip():
                    return nested
    return ""


def _extract_answer(record: Dict[str, Any]) -> str:
    for key in ("answer", "target", "gold", "gt_content"):
        value = record.get(key)
        if isinstance(value, str) and value.strip():
            return value
    return ""


def _select_last_boxed(text: str) -> str:
    boxed = BOXED_RE.findall(text or "")
    if boxed:
        return boxed[-1].strip()
    tag = TAG_RE.findall(text or "")
    if tag:
        last = tag[-1].strip()
        boxed = BOXED_RE.findall(last)
        if boxed:
            return boxed[-1].strip()
        return last
    return (text or "").strip()


def score_record(record: Dict[str, Any], val_only: bool = False) -> Optional[Dict[str, Any]]:
    if not isinstance(record, dict):
        return None

    record_id = str(record.get("id", ""))
    if val_only and not record_id.startswith("val"):
        return None

    prediction = _extract_prediction(record)
    answer = _extract_answer(record)
    scored = dict(record)
    score = 0
    exact = False

    if prediction and answer:
        last_boxed = _select_last_boxed(prediction)
        exact = last_boxed == answer.strip() or _normalize_text(last_boxed) == _normalize_text(answer)
        if exact:
            score = 1

    if not exact and int(record.get("hit", 0) or 0) == 1:
        score = 1

    scored["score"] = score
    scored["is_correct"] = bool(score)
    if prediction:
        scored["normalized_prediction"] = _select_last_boxed(prediction)
    return scored


def _load_json_records(path: Path) -> List[Dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        return [r for r in data if isinstance(r, dict)]
    if isinstance(data, dict):
        return [v for v in data.values() if isinstance(v, dict)]
    raise ValueError(f"unsupported JSON structure in {path}")


def _load_jsonl_records(path: Path) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    for line_num, line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        stripped = line.strip()
        if not stripped:
            continue
        try:
            record = json.loads(stripped)
        except json.JSONDecodeError as exc:
            records.append({"line_num": line_num, "parse_error": f"invalid JSON: {exc}", "raw_line": line})
            continue
        if isinstance(record, dict):
            records.append(record)
    return records


def _load_xlsx_records(path: Path, sheet: Optional[str] = None) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read XLSX input")
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = ["" if cell is None else str(cell) for cell in rows[0]]
    records: List[Dict[str, Any]] = []
    for row in rows[1:]:
        record = {}
        for header, value in zip(headers, row):
            if header:
                record[header] = value
        if record:
            records.append(record)
    return records


def load_records(path: str, sheet: Optional[str] = None) -> List[Dict[str, Any]]:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".jsonl":
        return _load_jsonl_records(source)
    if suffix == ".json":
        return _load_json_records(source)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx_records(source, sheet=sheet)
    raise ValueError(f"unsupported input type: {suffix}")


def write_records(path: str, records: Sequence[Dict[str, Any]]) -> None:
    output = Path(path)
    if output.suffix.lower() == ".json":
        output.write_text(json.dumps(list(records), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        return
    with output.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False) + "\n")


def summarize(records: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
    total = len(records)
    scored = [r for r in records if isinstance(r, dict) and "score" in r]
    correct = sum(1 for r in scored if int(r.get("score", 0) or 0) == 1)
    val_total = sum(1 for r in scored if str(r.get("id", "")).startswith("val"))
    val_correct = sum(
        1
        for r in scored
        if str(r.get("id", "")).startswith("val") and int(r.get("score", 0) or 0) == 1
    )
    return {
        "total": total,
        "scored": len(scored),
        "correct": correct,
        "accuracy": (correct / len(scored)) if scored else 0.0,
        "val_total": val_total,
        "val_correct": val_correct,
        "val_accuracy": (val_correct / val_total) if val_total else 0.0,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Score Skywork evaluation records with the last boxed answer.")
    parser.add_argument("--input", required=True, help="Input JSON, JSONL, or XLSX file.")
    parser.add_argument("--output", default="", help="Optional output file for annotated records.")
    parser.add_argument("--sheet", default="", help="Optional XLSX sheet name.")
    parser.add_argument("--val-only", action="store_true", help="Only score records whose id starts with val.")
    parser.add_argument("--json", action="store_true", help="Print summary as JSON.")
    args = parser.parse_args()

    records = load_records(args.input, sheet=args.sheet or None)
    scored_records = []
    for record in records:
        scored = score_record(record, val_only=args.val_only)
        if scored is not None:
            scored_records.append(scored)

    summary = summarize(scored_records)
    if args.output:
        write_records(args.output, scored_records)

    if args.json or args.output:
        print(json.dumps(summary, ensure_ascii=False, indent=2))
    else:
        print(f"Total: {summary['total']}")
        print(f"Scored: {summary['scored']}")
        print(f"Correct: {summary['correct']}")
        print(f"Accuracy: {summary['accuracy']:.4f}")
        print(f"Val total: {summary['val_total']}")
        print(f"Val correct: {summary['val_correct']}")
        print(f"Val accuracy: {summary['val_accuracy']:.4f}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
